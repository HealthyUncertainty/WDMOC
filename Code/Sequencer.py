# -*- coding: utf-8 -*-
"""
The Sequencer is the master file for the model. It creates entities and runs them through the model based
on their current model state, characteristics, history, and underlying probabilistic distrubtions of the 
model parameters.

The sequencer will run "num_entities" entities through the model and output the resources and events experienced
by each entity as a series of lists.

@author: icromwell
"""
############################################################################################
############################################################################################
# LOAD SOME NECESSARY PACKAGES AND FUNCTIONS

import time
import pickle
from openpyxl import load_workbook                  # Load the import function
import numpy

# Import Parameter Estimates from the table"

from Glb_Estimates import Estimates
from Glb_Estimates import Estimate

inbook = load_workbook('InputParameters.xlsx')
sheet = inbook["Inputs"]
estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))

del(estimates.Parameter)

with open('estimates.pickle', 'wb') as inputs:
    pickle.dump(estimates, inputs, pickle.HIGHEST_PROTOCOL)

# Import Regression Coefficients from the table

regsheet = inbook["RegCoeffs"]
source = []                                           # A list to hold data

"Convert the openpyxl object into a useable form"

for row in list(regsheet.rows)[1:]:
    args = [cell.value for cell in row]
    source.append(args)
    
for row in range(len(source)):
    source[row][0] = str(source[row][0])
    source[row][1] = str(source[row][1])

"Create a multi-level dictionary to hold each parameter from the regression model:"
    
config = {}         # creates the blank dictionary
for param, factor, vartype, level, mean, SE in source:
    SE = SE if SE else 0    # If SE is blank, enter zero
    vartype = vartype if vartype else 0
    mean = mean if mean not in ("ref", None) else 0     # Reference category = 0
    if param not in config:
        config[param] = {}
    
    if level:
        if factor not in config[param]:
            config[param][factor] = {"vartype": vartype}
        config[param][factor][level] = {"mean": mean, "SE": SE}
    else:
        config[param][factor] = {"vartype": vartype, "mean": mean, "SE": SE}

with open('regcoeffs.pickle', 'wb') as regcoeffs:
    pickle.dump(config, regcoeffs, pickle.HIGHEST_PROTOCOL)
    
# Import cost estimates from the table

costsheet = inbook["Costs"]
cost_estimates = Estimates()
for line in costsheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(cost_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(cost_estimates.Parameter)

# Create a dictionary of unit costs that the program will read from
CostDict = {}

for i in range(0, costsheet.max_row):
    cost_name = str(costsheet.cell(row = i+1, column = 1).value)
    cost_type = costsheet.cell(row = i+1, column = 2).value
    cost_mean = costsheet.cell(row = i+1, column = 3).value
    cost_se = costsheet.cell(row = i+1, column = 4).value
    CostDict[cost_name] = (cost_type, cost_mean, cost_se)
del(CostDict['Parameter'])
#del(CostDict['None'])

#############################################################################################
############################################################################################

################################
# STEP 1 - SET UP THE SEQUENCER
"Define the number of entities you want to model"

num_entities = 50000

ResourceList = []
EventsList = []
QALYList = []
EntityList = []

with open('estimates.pickle', 'rb') as f:
    estimates = pickle.load(f)

with open('regcoeffs.pickle', 'rb') as f:
    regcoeffs = pickle.load(f)

################################
# STEP 2 - RUN THE SEQUENCER
looptime_start = time.time()
for i in range(0, num_entities):
    entity_num = 'entity' + str(i + 1)                  # Create an entity for each iteration of the model
    
    from Glb_CreateEntity import Entity
    entity_num = Entity()
    entity = entity_num                             # Identify the current entity as the most recently created one
    #print("Entity %2.0f is created"%(i+1))
    if i % (num_entities/20) == 0:
        print("Now simulating entity", i, "of", num_entities)
      
    "Create resource table"
    
    resources = []
    events = []
    natHist = []
    QALY = []

    while True:
        
        #Apply Demographic Characteristics and Natural History to a newly-created entity
    
        if entity.stateNum == 0.0:
            from Glb_ApplyInit import ApplyInit
            applyinit = ApplyInit(estimates)
            applyinit.Process(entity)
            from NatHist_OralCancer import NatHistOCa
            nathistoca = NatHistOCa(estimates, regcoeffs)
            nathistoca.Process(entity, natHist)
            
        #PROBABILITY NODE: Does this person regularly see a dentist?
            # If yes - develop OPL while undergoing regular observations (state 1.0)
            # If no - develop OPL and possibly cancer (state 1.8)
            
        if entity.stateNum == 0.1:
            if entity.hasDentist == 1:
                entity.stateNum = 1.0
                entity.currentState = "1.0 - Undergoing regular dental screening"
            elif entity.hasDentist == 0:
                entity.stateNum = 1.8
                entity.currentState = "1.8 - No access to dentist"
            else:
                print("The entity has not been assigned a value for 'hasDentist'. The simulation must terminate")
                entity.stateNum = 0.9
            
        ### Advance the clock to next scheduled event (NatHist, Sysp, Recurrence, Death) ###
    
        from Glb_CheckTime import CheckTime
        CheckTime(entity, estimates, natHist, QALY)

        ### Run next scheduled event/process according to state ###
    
        #People with a participating dentist undergo regular screening appointments    
        if entity.stateNum == 1.0:                
            from SysP_ScreenAppt import ScreenAppt
            screenappt = ScreenAppt(estimates, regcoeffs)            
            screenappt.Process(entity)
           
        #People with no dentist wait for disease event        
        if entity.stateNum == 1.8:
            entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
                
        #People with a detected premalignancy undergo regular follow-up        
        if entity.stateNum == 2.0:
            from SysP_OPLmanage import OPLManage
            oplmanage = OPLManage(estimates, regcoeffs)            
            oplmanage.Process(entity) 
        
        #People with a detected cancer undergo treatment       
        if entity.stateNum == 3.0:
            from SysP_IncidentCancer import IncidentCancer
            incidentcancer = IncidentCancer(estimates, regcoeffs)
            incidentcancer.Process(entity)        

        #People who have been successfully treated undergo regular follow-up     
        if entity.stateNum == 4.0:
            from SysP_Followup import Followup
            followup = Followup(estimates, regcoeffs)
            followup.Process(entity)
   
        #People whose disease has entered remission after 10 years     
        if entity.stateNum == 4.8:
            #entity is in remission, no further events occur
            entity.allTime = entity.natHist_deathAge + 0.0001
             
        #People with terminal disease receive palliative care     
        if entity.stateNum == 5.0:
            from SysP_Terminal import Terminal
            terminal = Terminal(estimates, regcoeffs)
            terminal.Process(entity)
      
        #The entity is dead      
        if entity.stateNum == 100:
            #print("Entity is", entity.death_desc, "at:", entity.time_death)
            events.append(('Entity dies', entity.time_death))
            entity.utility.append(('Dead', 0, entity.time_death))
            break
        
        # An error has occurred
        if entity.stateNum == 99:
            print("An error has occurred and the simulation must end")
            print(entity.currentState)
            break
        
    EntityList.append(entity)
    ResourceList.append(entity.resources)
    QALYList.append(entity.utility)
    
    # END WHILE
    
looptime_end = time.time()
looptime = round((looptime_end - looptime_start)/60, 2)
print("The sequencer simulated", num_entities, "entities. It took", looptime, "minutes.")

################################
# OPTIONAL STEP - SAVE OUTPUTS TO DISK 

numpy.save('EntityList', EntityList)
numpy.save('ResourceList', ResourceList)
numpy.save('EventsList', EventsList)
numpy.save('QALYList', QALYList)

################################
# STEP 3 - ESTIMATE COSTS AND SURVIVAL FOR SIMULATED COHORT
from Glb_AnalyzeOutput import Analyze_Output
output = Analyze_Output(estimates, CostDict)

# Estimate the costs generated by the entities in the population
CohortCost = []
for i in range(len(EntityList)):
    entity = EntityList[i]
    CohortCost.append(output.EntityCost(entity))
    
# Estimate the LYG and QALY generated by the entities in the population
CohortSurvival = np.array([output.EntitySurvival(x) for x in EntityList])
CohortLYG = CohortSurvival[:,0]
CohortQALY = CohortSurvival[:,1]


####################################################
# FOOTNOTE:
#
#   1 - The clock moves forward an arbitrary number of days, but is reset to the next natural
#       history or disease event by 'Glb_Checktime.py'. The purpose of moving the clock forward is
#       simply to prompt advancement to the next event.
